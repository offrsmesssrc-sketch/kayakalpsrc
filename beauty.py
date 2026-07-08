#!/usr/bin/env python3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os, sys, json, re, shutil, socket, threading, base64
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk
import qrcode
import ssl
from http.server import BaseHTTPRequestHandler, HTTPServer, ThreadingHTTPServer

# Face recognition imports
import cv2
import face_recognition
import numpy as np

# ----------- Paths and constants -----------
DEFAULT_DATA_FOLDER = "D:/beautyparlour"
if not os.path.exists(DEFAULT_DATA_FOLDER):
    try:
        os.makedirs(DEFAULT_DATA_FOLDER)
    except:
        pass

if os.path.exists(DEFAULT_DATA_FOLDER):
    DATA_FOLDER = DEFAULT_DATA_FOLDER
else:
    fallback_paths = [
        "C:/beautyparlour",
        os.path.join(os.path.expanduser("~"), "AppData", "Local", "beautyparlour"),
        os.path.join(os.path.expanduser("~"), "beautyparlour"),
    ]
    DATA_FOLDER = None
    for path in fallback_paths:
        try:
            if not os.path.exists(path):
                os.makedirs(path)
            DATA_FOLDER = path
            break
        except Exception:
            continue
    if not DATA_FOLDER:
        DATA_FOLDER = os.path.join(os.getcwd(), "beautyparlour")
        if not os.path.exists(DATA_FOLDER):
            os.makedirs(DATA_FOLDER)

FILE_NAME = os.path.join(DATA_FOLDER, "pers_data.xlsx")
SHEET_NAME = "pers"
ATTENDANCE_FILE = os.path.join(DATA_FOLDER, "attendance.xlsx")
SERVICES_LOG_FILE = os.path.join(DATA_FOLDER, "services_log.xlsx")

COLUMNS = ["ID", "Name"]

# ----------- Face Recognition Configuration -----------
TOLERANCE = 0.48
MODEL = 'hog'
KNOWN_FACES_DIR = os.path.join(DATA_FOLDER, "faces_kaya")
ENCODINGS_CACHE_FILE = os.path.join(DATA_FOLDER, "face_encodings_cache.json")

known_encodings = []
known_ids = []
known_names = {}

def load_known_faces():
    global known_encodings, known_ids, known_names
    known_encodings = []
    known_ids = []
    known_names = {}
    
    ensure_workbook()
    try:
        wb = load_workbook(FILE_NAME, read_only=True)
        ws = wb[SHEET_NAME]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                known_names[str(row[0]).strip()] = str(row[1] or "").strip()
        wb.close()
    except Exception as e:
        print("Error loading names for face recognition:", e)
        
    if not os.path.exists(KNOWN_FACES_DIR):
        os.makedirs(KNOWN_FACES_DIR)
        return
        
    cache = {}
    if os.path.exists(ENCODINGS_CACHE_FILE):
        try:
            with open(ENCODINGS_CACHE_FILE, "r") as f:
                cache = json.load(f)
        except Exception:
            pass
            
    updated_cache = {}
    valid_extensions = (".jpg", ".jpeg", ".png")
    
    for filename in os.listdir(KNOWN_FACES_DIR):
        if not filename.lower().endswith(valid_extensions):
            continue
        filepath = os.path.join(KNOWN_FACES_DIR, filename)
        b_id = os.path.splitext(filename)[0].strip()
        mtime = os.path.getmtime(filepath)
        
        cached_item = cache.get(b_id)
        encoding = None
        if cached_item and cached_item.get("mtime") == mtime:
            encoding = cached_item.get("encoding")
            
        if encoding is None:
            try:
                img = face_recognition.load_image_file(filepath)
                encs = face_recognition.face_encodings(img)
                if encs:
                    encoding = list(encs[0])
                    print(f"Face encoded for Beautician ID: {b_id}")
                else:
                    print(f"No face found in image for ID: {b_id}")
            except Exception as e:
                print(f"Error encoding face for {b_id}: {e}")
                
        if encoding:
            known_encodings.append(np.array(encoding))
            known_ids.append(b_id)
            updated_cache[b_id] = {
                "mtime": mtime,
                "encoding": encoding
            }
            
    try:
        with open(ENCODINGS_CACHE_FILE, "w") as f:
            json.dump(updated_cache, f)
    except Exception as e:
        print("Failed to save face cache:", e)

# ----------- Login settings -----------
LOGIN_USERNAME = "admin"
LOGIN_PASSWORD = "abcd@1234"
WEB_CRED_FILE = os.path.join(DATA_FOLDER, "web_credentials.json")

def load_web_credentials():
    try:
        if os.path.exists(WEB_CRED_FILE):
            with open(WEB_CRED_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data.get("username", "admin"), data.get("password", "abcd@1234")
    except Exception:
        pass
    return "admin", "abcd@1234"

def save_web_credentials(username, password):
    try:
        if not os.path.exists(DATA_FOLDER):
            os.makedirs(DATA_FOLDER)
        with open(WEB_CRED_FILE, "w", encoding="utf-8") as f:
            json.dump({"username": username, "password": password}, f)
    except Exception as e:
        print(f"Failed to save web credentials: {e}")

# ----------- Rate List (from rate.pptx) -----------
RATE_LIST = {
    "Threading": [
        {"name": "Eye Brow", "rate": 30},
        {"name": "Up Lips", "rate": 20},
        {"name": "Fore Head", "rate": 20},
        {"name": "Chin", "rate": 10},
        {"name": "Face Side", "rate": 50},
        {"name": "Full Side", "rate": 100}
    ],
    "Face Waxing": [
        {"name": "Forehead Wax", "rate": 60},
        {"name": "Chin Wax", "rate": 45},
        {"name": "Face Side Wax", "rate": 150},
        {"name": "Full Face Wax", "rate": 300},
        {"name": "Up Lips Wax", "rate": 50}
    ],
    "Bleach & De-Tan": [
        {"name": "Fruit Bleach", "rate": 120},
        {"name": "Oxy Bleach", "rate": 130},
        {"name": "Olivia Bleach", "rate": 120},
        {"name": "Nature Gold Bleach", "rate": 160},
        {"name": "Nature Pro Lato Bleach", "rate": 200},
        {"name": "Vedic Line Charcoal Bleach", "rate": 130},
        {"name": "De Tan Nature Pack", "rate": 300},
        {"name": "O3++ De Tan Pack", "rate": 500}
    ],
    "Facials": [
        {"name": "Gold Facial Nature", "rate": 500},
        {"name": "Lotus with Glow", "rate": 600},
        {"name": "Papaya Fruit Facial", "rate": 600},
        {"name": "VLCC Pearl n Diamond", "rate": 550},
        {"name": "VLCC Instant Glow", "rate": 500},
        {"name": "Nutri Glow", "rate": "Ask"},
        {"name": "Shahnaz Facial (Gold)", "rate": 1500},
        {"name": "Lotus Gold Facial", "rate": "Ask"},
        {"name": "De-tan Facial", "rate": 500},
        {"name": "Wedding Glow", "rate": "Ask"},
        {"name": "Bridal Facial", "rate": "Ask"},
        {"name": "Ozone Facial / Xpress Facial", "rate": 700},
        {"name": "O3++ Facial", "rate": 2000},
        {"name": "Korean Facial (Glass)", "rate": 700},
        {"name": "Aroma Facial (Aloe-vera)", "rate": 600},
        {"name": "Vedic Facial (Vit C)", "rate": 600},
        {"name": "Thermo Herb Facial", "rate": 1000},
        {"name": "Home Remedies - Fruit Original Juice Facial", "rate": 500}
    ],
    "Body Waxing": [
        {"name": "Full Arms - White Chocolate", "rate": 200},
        {"name": "Full Arms - Aloe Vera", "rate": 200},
        {"name": "Full Arms - Rica", "rate": 400},
        {"name": "Full Legs - White Chocolate", "rate": 500},
        {"name": "Full Legs - Aloe Vera", "rate": 500},
        {"name": "Full Legs - Rica", "rate": 600},
        {"name": "Half Legs - White Chocolate", "rate": 250},
        {"name": "Half Legs - Aloe Vera", "rate": 250},
        {"name": "Half Legs - Rica", "rate": 300},
        {"name": "Under Arms - White Chocolate", "rate": 80},
        {"name": "Under Arms - Aloe Vera", "rate": 80},
        {"name": "Under Arms - Rica", "rate": 100}
    ],
    "Hair Work": [
        {"name": "Smoothening / Straightening (Glatt / Strax)", "rate": "Ask"},
        {"name": "Botox (Kera BTX+)", "rate": "Ask"},
        {"name": "Hair Highlights (Streax)", "rate": "Ask"},
        {"name": "Root Touchup", "rate": 500},
        {"name": "Hair Global - Short", "rate": 1200},
        {"name": "Hair Global - Long", "rate": 1800},
        {"name": "Hair Spa (Streax) - Short", "rate": 500},
        {"name": "Hair Spa (Streax) - Long", "rate": 550},
        {"name": "Hair Spa (Loreal) - Short", "rate": 700},
        {"name": "Hair Spa (Loreal) - Long", "rate": 750},
        {"name": "Keratin Kerafine", "rate": 3000}
    ],
    "Body Massage": [
        {"name": "Full Body Oil Massage - Baby/Olive/Coconut", "rate": 1000},
        {"name": "Full Body Cream Massage", "rate": 1100},
        {"name": "Foot Massage - Oil / Cream", "rate": 250}
    ],
    "Hair Cut": [
        {"name": "Straight Cut", "rate": 100},
        {"name": "U-Cut", "rate": 120},
        {"name": "Deep U Cut", "rate": 150},
        {"name": "Front Layer Cut", "rate": 150},
        {"name": "Three Step Cut", "rate": 300},
        {"name": "Full Layer Cut", "rate": 350},
        {"name": "Multi Step Cut", "rate": 400},
        {"name": "Blunt Cut", "rate": 300},
        {"name": "Bob Hair Cut", "rate": 300},
        {"name": "Butterfly Hair Cut", "rate": 400}
    ],
    "Hand & Feet Care": [
        {"name": "Pedicure (Fruit / VLCC)", "rate": 350},
        {"name": "Manicure (Fruit / VLCC)", "rate": 350},
        {"name": "Pedicure + Bleach", "rate": 550},
        {"name": "Manicure + Bleach", "rate": 550},
        {"name": "Pedicure (Paraffin)", "rate": "Ask"},
        {"name": "Manicure (Paraffin)", "rate": "Ask"}
    ],
    "Other Work": [
        {"name": "Blow Dry", "rate": 100},
        {"name": "Hair Wash", "rate": 150},
        {"name": "Hair Oil Head", "rate": 250},
        {"name": "Ironing + Setting Spray", "rate": 250},
        {"name": "Mehandi (Head)", "rate": 200},
        {"name": "Hair Style", "rate": "Ask"},
        {"name": "Mehandi Hand - Half", "rate": 200},
        {"name": "Mehandi Hand - Full", "rate": 300},
        {"name": "Party Makeup (PAC / Mrs)", "rate": 1000},
        {"name": "HD Makeup (Forever 52 / PAC / MAC)", "rate": 1500},
        {"name": "Dress Draping", "rate": 200},
        {"name": "Nail Extensions / Art", "rate": "Ask"}
    ],
    "Special Offers": [
        {"name": "Glow Starter Combo - Eye Brow + Up Lips + Fore Head + Chin", "rate": 75},
        {"name": "Full Grooming Combo - Full Face Wax + Eye Brow", "rate": 320},
        {"name": "Hand & Feet Combo - Manicure + Pedicure", "rate": 650},
        {"name": "Bridal Glow Consultation", "rate": 0},
        {"name": "Hair Spa Day - Streax Hair Spa", "rate": 499},
        {"name": "Party Ready Add-on - Dress Draping + Blow Dry", "rate": 280},
        {"name": "Premium Facial Upgrade", "rate": "Ask"},
        {"name": "Student / Trainee Grooming Day", "rate": "Ask"}
    ]
}

RATES_FILE = os.path.join(DATA_FOLDER, "rates.json")

def load_rates():
    global RATE_LIST
    if os.path.exists(RATES_FILE):
        try:
            with open(RATES_FILE, "r", encoding="utf-8") as f:
                RATE_LIST = json.load(f)
        except Exception as e:
            print("Error loading rates.json:", e)

def save_rates():
    try:
        with open(RATES_FILE, "w", encoding="utf-8") as f:
            json.dump(RATE_LIST, f, indent=4)
    except Exception as e:
        print("Error saving rates.json:", e)

load_rates()

# ----------- Excel Helpers -----------
def ensure_workbook():
    if not os.path.exists(DATA_FOLDER):
        os.makedirs(DATA_FOLDER)
    if os.path.exists(FILE_NAME):
        try:
            wb = load_workbook(FILE_NAME)
            if SHEET_NAME not in wb.sheetnames:
                ws = wb.active
                ws.title = SHEET_NAME
            ws = wb[SHEET_NAME]
            
            headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
            if headers != COLUMNS:
                rows_data = []
                for r in range(2, ws.max_row + 1):
                    id_val = ws.cell(row=r, column=1).value
                    name_val = ws.cell(row=r, column=2).value
                    if id_val or name_val:
                        rows_data.append((str(id_val or "").strip(), str(name_val or "").strip()))
                
                wb.remove(ws)
                ws = wb.create_sheet(SHEET_NAME)
                ws.cell(row=1, column=1, value="ID")
                ws.cell(row=1, column=2, value="Name")
                for idx, (b_id, name) in enumerate(rows_data, start=2):
                    ws.cell(row=idx, column=1, value=b_id)
                    ws.cell(row=idx, column=2, value=name)
            wb.save(FILE_NAME)
            wb.close()
        except Exception as e:
            print("Error validating pers_data:", e)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.cell(row=1, column=1, value="ID")
        ws.cell(row=1, column=2, value="Name")
        wb.save(FILE_NAME)
        wb.close()

def ensure_attendance_workbook():
    headers = ["S.no", "Date", "Beautician ID", "Name", "Check-in Time", "Check-out Time"]
    if os.path.exists(ATTENDANCE_FILE):
        try:
            wb = load_workbook(ATTENDANCE_FILE)
            ws = wb.active
            current_headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
            if current_headers != headers:
                rows_data = []
                for r in range(2, ws.max_row + 1):
                    row_vals = [ws.cell(row=r, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
                    if len(row_vals) >= 8:
                        b_id = str(row_vals[1] or "").strip()
                        name = str(row_vals[3] or "").strip()
                        date_str = str(row_vals[6] or "").strip()
                        time_str = str(row_vals[7] or "").strip()
                        rows_data.append((date_str, b_id, name, time_str, ""))
                    elif len(row_vals) >= 5:
                        b_id = str(row_vals[1] or "").strip()
                        name = str(row_vals[2] or "").strip()
                        date_str = str(row_vals[3] or "").strip()
                        time_str = str(row_vals[4] or "").strip()
                        rows_data.append((date_str, b_id, name, time_str, ""))
                
                wb.remove(ws)
                ws = wb.create_sheet("Sheet")
                for i, h in enumerate(headers, start=1):
                    ws.cell(row=1, column=i, value=h)
                for idx, (date_str, b_id, name, check_in, check_out) in enumerate(rows_data, start=2):
                    ws.cell(row=idx, column=1, value=idx-1)
                    ws.cell(row=idx, column=2, value=date_str)
                    ws.cell(row=idx, column=3, value=b_id)
                    ws.cell(row=idx, column=4, value=name)
                    ws.cell(row=idx, column=5, value=check_in)
                    ws.cell(row=idx, column=6, value=check_out)
            wb.save(ATTENDANCE_FILE)
            wb.close()
        except Exception as e:
            print("Error validating attendance file:", e)
    else:
        wb = Workbook()
        ws = wb.active
        for i, h in enumerate(headers, start=1):
            ws.cell(row=1, column=i, value=h)
        wb.save(ATTENDANCE_FILE)
        wb.close()

def ensure_services_log_workbook():
    headers = ["S.no", "Date", "Time", "Beautician ID", "Beautician Name", "Category", "Service Name", "Rate"]
    if os.path.exists(SERVICES_LOG_FILE):
        try:
            wb = load_workbook(SERVICES_LOG_FILE)
            ws = wb.active
            current_headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
            if not current_headers or current_headers[0] == "":
                for i, h in enumerate(headers, start=1):
                    ws.cell(row=1, column=i, value=h)
                wb.save(SERVICES_LOG_FILE)
            wb.close()
        except Exception:
            pass
    else:
        wb = Workbook()
        ws = wb.active
        for i, h in enumerate(headers, start=1):
            ws.cell(row=1, column=i, value=h)
        wb.save(SERVICES_LOG_FILE)
        wb.close()

def find_row_by_army_no(ws, id_val):
    target = str(id_val).strip()
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip() == target:
            return r
    return None

def compute_beautician_stats(beautician_id):
    today_str = datetime.now().strftime("%d/%m/%Y")
    current_month_year = datetime.now().strftime("%m/%Y")
    
    services_today = 0
    services_month = 0
    services_total = 0
    money_today = 0
    money_month = 0
    money_total = 0
    
    if os.path.exists(SERVICES_LOG_FILE):
        try:
            wb = load_workbook(SERVICES_LOG_FILE, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 8:
                    continue
                date_val = str(row[1] or "").strip()
                id_val = str(row[3] or "").strip()
                rate_val = 0
                try:
                    rate_val = float(row[7]) if row[7] is not None else 0
                except ValueError:
                    pass
                
                if id_val == str(beautician_id).strip():
                    services_total += 1
                    money_total += rate_val
                    
                    if date_val == today_str:
                        services_today += 1
                        money_today += rate_val
                    
                    if len(date_val) == 10:
                        m_y = date_val[3:]
                        if m_y == current_month_year:
                            services_month += 1
                            money_month += rate_val
            wb.close()
        except Exception as e:
            print("Error computing beautician stats:", e)
            
    return {
        "services_today": services_today,
        "services_month": services_month,
        "services_total": services_total,
        "money_today": int(money_today),
        "money_month": int(money_month),
        "money_total": int(money_total)
    }

def compute_overall_stats():
    today_str = datetime.now().strftime("%d/%m/%Y")
    monthly_earnings = {}
    total_money_today = 0
    
    if os.path.exists(SERVICES_LOG_FILE):
        try:
            wb = load_workbook(SERVICES_LOG_FILE, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 8:
                    continue
                date_val = str(row[1] or "").strip()
                rate_val = 0
                try:
                    rate_val = float(row[7]) if row[7] is not None else 0
                except ValueError:
                    pass
                
                if date_val == today_str:
                    total_money_today += rate_val
                
                if len(date_val) == 10:
                    m_y = date_val[3:] # MM/YYYY
                    monthly_earnings[m_y] = monthly_earnings.get(m_y, 0) + rate_val
            wb.close()
        except Exception as e:
            print("Error computing overall stats:", e)
            
    formatted_monthly = []
    try:
        sorted_months = sorted(monthly_earnings.keys(), key=lambda x: datetime.strptime(x, "%m/%Y"), reverse=True)
        for m_y in sorted_months:
            dt = datetime.strptime(m_y, "%m/%Y")
            month_name = dt.strftime("%B %Y")
            formatted_monthly.append({
                "month_str": month_name,
                "amount": int(monthly_earnings[m_y])
            })
    except Exception as e:
        print("Sorting monthly logs failed:", e)
        
    return {
        "total_today": int(total_money_today),
        "monthly_breakdown": formatted_monthly
    }

def get_beauticians_list():
    ensure_workbook()
    beauticians = []
    try:
        wb = load_workbook(FILE_NAME)
        ws = wb[SHEET_NAME]
        for r in range(2, ws.max_row + 1):
            id_val = ws.cell(row=r, column=1).value
            name_val = ws.cell(row=r, column=2).value
            if id_val:
                b_id = str(id_val).strip()
                name = str(name_val or "").strip()
                stats = compute_beautician_stats(b_id)
                beauticians.append({
                    "id": b_id,
                    "name": name,
                    "stats": stats
                })
        wb.close()
    except Exception as e:
        print("Error getting beauticians list:", e)
    return beauticians

def log_beautician_attendance_face(beautician_id, name):
    ensure_attendance_workbook()
    today_str = datetime.now().strftime("%d/%m/%Y")
    now_time = datetime.now().strftime("%H:%M:%S")
    
    try:
        wb = load_workbook(ATTENDANCE_FILE)
        ws = wb.active
        
        # Find the last record of this beautician today
        last_row = None
        for r in range(2, ws.max_row + 1):
            date_val = str(ws.cell(row=r, column=2).value or "").strip()
            id_val = str(ws.cell(row=r, column=3).value or "").strip()
            if date_val == today_str and id_val == str(beautician_id).strip():
                last_row = r
                
        if last_row is None:
            # First check-in of the day
            next_sno = 1
            max_r = ws.max_row
            if max_r > 1:
                last_val = ws.cell(row=max_r, column=1).value
                try:
                    next_sno = int(last_val) + 1
                except:
                    next_sno = max_r
            ws.append([next_sno, today_str, beautician_id, name, now_time, ""])
            action = "Checked In"
            time_logged = now_time
        else:
            # Check if last session is already checked out
            check_in_str = str(ws.cell(row=last_row, column=5).value or "").strip()
            check_out_str = str(ws.cell(row=last_row, column=6).value or "").strip()
            
            if check_out_str:
                # Last session checked out, this is a new check-in
                next_sno = 1
                max_r = ws.max_row
                if max_r > 1:
                    last_val = ws.cell(row=max_r, column=1).value
                    try:
                        next_sno = int(last_val) + 1
                    except:
                        next_sno = max_r
                ws.append([next_sno, today_str, beautician_id, name, now_time, ""])
                action = "Checked In"
                time_logged = now_time
            else:
                # Attempting check-out - always allowed
                ws.cell(row=last_row, column=6, value=now_time)
                action = "Checked Out"
                time_logged = now_time
                
        wb.save(ATTENDANCE_FILE)
        wb.close()
        return True, action, time_logged
    except Exception as e:
        print("Error logging face attendance:", e)
        return False, "Database error", ""

def log_session_transactions(beautician_id, name, services_list):
    ensure_services_log_workbook()
    if not services_list:
        return True
    try:
        wb = load_workbook(SERVICES_LOG_FILE)
        ws = wb.active
        
        next_sno = 1
        max_r = ws.max_row
        if max_r > 1:
            last_val = ws.cell(row=max_r, column=1).value
            try:
                next_sno = int(last_val) + 1
            except:
                next_sno = max_r
                
        now = datetime.now()
        date_str = now.strftime("%d/%m/%Y")
        time_str = now.strftime("%H:%M:%S")
        
        for item in services_list:
            category = item.get("category", "")
            service_name = item.get("name", "")
            rate = 0
            try:
                rate = float(item.get("rate", 0))
            except:
                pass
            
            ws.append([
                next_sno,
                date_str,
                time_str,
                beautician_id,
                name,
                category,
                service_name,
                rate
            ])
            next_sno += 1
            
        wb.save(SERVICES_LOG_FILE)
        wb.close()
        return True
    except Exception as e:
        print("Error logging session transactions:", e)
        return False

def log_service_transaction(beautician_id, name, category, service_name, rate):
    ensure_services_log_workbook()
    try:
        wb = load_workbook(SERVICES_LOG_FILE)
        ws = wb.active
        next_sno = 1
        max_r = ws.max_row
        if max_r > 1:
            last_val = ws.cell(row=max_r, column=1).value
            try:
                next_sno = int(last_val) + 1
            except:
                next_sno = max_r
        now = datetime.now()
        ws.append([
            next_sno,
            now.strftime("%d/%m/%Y"),
            now.strftime("%H:%M:%S"),
            beautician_id,
            name,
            category,
            service_name,
            rate
        ])
        wb.save(SERVICES_LOG_FILE)
        wb.close()
        return True
    except Exception as e:
        print("Error logging service transaction:", e)
        return False

# ----------- Web Server Handler & Template -----------
class BeautyHandler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        pass

    def do_GET(self):
        if self.path == '/':
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(HTML_TEMPLATE.encode('utf-8'))
        elif self.path == '/api/get_rate_list':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(RATE_LIST).encode('utf-8'))
        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length)
        params = {}
        if post_data:
            try:
                params = json.loads(post_data.decode('utf-8'))
            except Exception:
                pass

        if self.path == '/api/get_dashboard':
            b_id = params.get("beautician_id")
            
            if b_id:
                beauticians = get_beauticians_list()
                b_id_str = str(b_id).strip()
                beauticians = [b for b in beauticians if b["id"] == b_id_str]
                # Return this beautician's personal earnings only
                personal = compute_beautician_stats(b_id_str)
                month_total = personal.get("money_month", 0)
                current_month_str = datetime.now().strftime("%B %Y")
                overall = {
                    "total_today": personal.get("money_today", 0),
                    "monthly_breakdown": [
                        {"month_str": current_month_str, "amount": month_total}
                    ] if month_total > 0 else []
                }
            else:
                beauticians = []
                overall = {
                    "total_today": 0,
                    "monthly_breakdown": []
                }
                
            res = {
                "success": True,
                "beauticians": beauticians,
                "overall": overall
            }
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(res).encode('utf-8'))

        elif self.path == '/api/get_rate_list':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(RATE_LIST).encode('utf-8'))

        elif self.path == '/api/log_attendance':
            b_id = params.get("beautician_id")
            name = params.get("name")
            success, action, time_logged = log_beautician_attendance_face(b_id, name)
            res = {"success": success, "action": action, "time": time_logged}
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(res).encode('utf-8'))

        elif self.path == '/api/log_service':
            b_id = params.get("beautician_id")
            name = params.get("beautician_name")
            service_name = params.get("service_name")
            category = params.get("category")
            rate = 0
            try:
                rate = float(params.get("rate", 0))
            except:
                pass
            
            success = log_service_transaction(b_id, name, category, service_name, rate)
            res = {"success": success}
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(res).encode('utf-8'))

        elif self.path == '/api/log_session':
            b_id = params.get("beautician_id")
            name = params.get("beautician_name")
            services = params.get("services", [])
            
            success = log_session_transactions(b_id, name, services)
            res = {"success": success}
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(res).encode('utf-8'))
            
        elif self.path == '/api/recognize':
            image_b64 = params.get("image", "")
            if not image_b64:
                res = {"success": False, "error": "No image data"}
            else:
                if "," in image_b64:
                    image_b64 = image_b64.split(",", 1)[1]
                try:
                    img_bytes = base64.b64decode(image_b64)
                    
                    load_known_faces()
                    
                    nparr = np.frombuffer(img_bytes, np.uint8)
                    frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
                    if frame is None:
                        res = {"success": False, "error": "Could not decode image"}
                    else:
                        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                        face_locs = face_recognition.face_locations(rgb, model=MODEL)
                        face_encs = face_recognition.face_encodings(rgb, face_locs)
                        
                        if not face_locs:
                           res = {"success": False, "error": "No face detected. Please align your face inside the camera view."}
                        elif not known_encodings:
                            res = {"success": False, "error": "No beautician photos saved on server. Please add photos in the desktop application."}
                        else:
                            enc = face_encs[0]
                            distances = face_recognition.face_distance(known_encodings, enc)
                            best_idx = np.argmin(distances) if len(distances) > 0 else None
                            
                            if best_idx is not None and distances[best_idx] < TOLERANCE:
                                matched_id = known_ids[best_idx]
                                matched_name = known_names.get(matched_id, "Unknown Beautician")
                                
                                success, action, time_logged = log_beautician_attendance_face(matched_id, matched_name)
                                if success:
                                    res = {
                                        "success": True,
                                        "name": matched_name,
                                        "id": matched_id,
                                        "action": action,
                                        "time": time_logged
                                    }
                                else:
                                    res = {"success": False, "error": action}
                            else:
                                res = {"success": False, "error": "Face not recognized. Please verify you are registered and your photo is uploaded."}
                except Exception as e:
                    res = {"success": False, "error": str(e)}
                    
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps(res).encode('utf-8'))
        else:
            self.send_response(404)
            self.end_headers()

# Premium Web Single Page Application Template
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Kaya Kalp Beauty Parlour</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <style>
        :root {
            --primary: #d48a97;
            --primary-hover: #c47683;
            --accent: #e5a9b4;
            --bg-dark: #12090b;
            --bg-light: #1c1114;
            --card-bg: rgba(43, 24, 30, 0.65);
            --card-border: rgba(229, 169, 180, 0.15);
            --card-hover-border: rgba(229, 169, 180, 0.4);
            --text-color: #fceef0;
            --text-sub: #d9c5c7;
            --accent-glow: rgba(229, 169, 180, 0.25);
        }
        * {
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }
        body {
            font-family: 'Outfit', sans-serif;
            background: radial-gradient(circle at top, var(--bg-light) 0%, var(--bg-dark) 100%);
            color: var(--text-color);
            min-height: 100vh;
            padding: 20px;
            display: flex;
            justify-content: center;
        }
        .container {
            width: 100%;
            max-width: 1200px;
        }
        header {
            text-align: center;
            margin-bottom: 30px;
            padding-bottom: 20px;
            border-bottom: 1px solid var(--card-border);
        }
        h1 {
            font-size: 34px;
            font-weight: 700;
            color: #ffffff;
            background: linear-gradient(135deg, #ffffff 0%, var(--accent) 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            text-shadow: 0 0 10px rgba(229, 169, 180, 0.1);
        }
        .subtitle {
            font-size: 13px;
            color: var(--text-sub);
            margin-top: 5px;
            letter-spacing: 2px;
            text-transform: uppercase;
        }
        
        /* Views */
        .view-panel {
            display: none;
        }
        .view-panel.active {
            display: block;
        }
        
        /* Dashboard View */
        .card {
            background: var(--card-bg);
            backdrop-filter: blur(12px);
            -webkit-backdrop-filter: blur(12px);
            border: 1px solid var(--card-border);
            border-radius: 16px;
            padding: 24px;
            box-shadow: 0 10px 30px rgba(0, 0, 0, 0.4);
            margin-bottom: 20px;
        }
        .table-responsive {
            overflow-x: auto;
            border-radius: 8px;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            text-align: left;
            white-space: nowrap;
        }
        th {
            padding: 16px;
            color: #ffffff;
            font-weight: 600;
            border-bottom: 2px solid var(--card-border);
            font-size: 13px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        td {
            padding: 16px;
            border-bottom: 1px solid var(--card-border);
            font-size: 15px;
            vertical-align: middle;
        }
        tr:hover td {
            background: rgba(229, 169, 180, 0.04);
        }
        
        /* Checkbox styling */
        .checkbox-container {
            display: inline-block;
            position: relative;
            cursor: pointer;
            width: 22px;
            height: 22px;
        }
        .checkbox-container input {
            opacity: 0;
            width: 0;
            height: 0;
        }
        .checkmark {
            position: absolute;
            top: 0;
            left: 0;
            height: 22px;
            width: 22px;
            background-color: rgba(255, 255, 255, 0.08);
            border: 1px solid var(--primary);
            border-radius: 6px;
            transition: all 0.2s;
        }
        .checkbox-container input:checked ~ .checkmark {
            background-color: var(--primary);
            border-color: var(--primary);
            box-shadow: 0 0 10px var(--primary);
        }
        .checkmark:after {
            content: "";
            position: absolute;
            display: none;
            left: 7px;
            top: 3px;
            width: 5px;
            height: 10px;
            border: solid white;
            border-width: 0 2px 2px 0;
            transform: rotate(45deg);
        }
        .checkbox-container input:checked ~ .checkmark:after {
            display: block;
        }
        
        .btn-proceed {
            padding: 14px;
            background: linear-gradient(135deg, var(--primary) 0%, var(--primary-hover) 100%);
            border: none;
            color: #ffffff;
            font-size: 16px;
            font-weight: 700;
            border-radius: 8px;
            cursor: pointer;
            transition: all 0.3s;
            box-shadow: 0 4px 12px rgba(212, 138, 151, 0.3);
            text-align: center;
            flex: 1;
            min-width: 220px;
        }
        .btn-proceed:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(212, 138, 151, 0.5);
        }
        
        /* Bottom Stats Cards */
        .bottom-stats {
            margin-top: 30px;
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
            gap: 20px;
        }
        .stat-card {
            background: var(--card-bg);
            border: 1px solid var(--card-border);
            border-radius: 16px;
            padding: 20px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.3);
        }
        .stat-title {
            font-size: 12px;
            color: var(--text-sub);
            text-transform: uppercase;
            letter-spacing: 1px;
            margin-bottom: 8px;
            font-weight: 600;
        }
        .stat-value {
            font-size: 26px;
            font-weight: 700;
            color: #ffffff;
        }
        .stat-list {
            max-height: 180px;
            overflow-y: auto;
            margin-top: 12px;
            padding-right: 5px;
        }
        .stat-item {
            display: flex;
            justify-content: space-between;
            padding: 8px 0;
            border-bottom: 1px solid rgba(229, 169, 180, 0.1);
            font-size: 14px;
        }
        .stat-item:last-child {
            border-bottom: none;
        }
        
        /* Rate List View */
        .rate-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 24px;
            flex-wrap: wrap;
            gap: 15px;
        }
        .beautician-badge {
            background: rgba(212, 138, 151, 0.15);
            border: 1px solid var(--primary);
            color: var(--accent);
            padding: 8px 18px;
            border-radius: 30px;
            font-weight: 600;
            font-size: 14px;
        }
        .category-tabs {
            display: flex;
            gap: 10px;
            overflow-x: auto;
            padding-bottom: 10px;
            margin-bottom: 25px;
            scrollbar-width: thin;
        }
        .category-tabs::-webkit-scrollbar {
            height: 6px;
        }
        .category-tabs::-webkit-scrollbar-thumb {
            background: rgba(229, 169, 180, 0.2);
            border-radius: 3px;
        }
        .category-tab {
            padding: 10px 20px;
            background: rgba(255, 255, 255, 0.04);
            border: 1px solid var(--card-border);
            border-radius: 30px;
            color: var(--text-sub);
            cursor: pointer;
            white-space: nowrap;
            font-weight: 500;
            font-size: 14px;
            transition: all 0.2s;
        }
        .category-tab:hover {
            border-color: var(--card-hover-border);
            color: #ffffff;
        }
        .category-tab.active {
            background: var(--primary);
            color: #ffffff;
            border-color: var(--primary);
            box-shadow: 0 0 10px rgba(212, 138, 151, 0.3);
        }
        .services-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(220px, 1fr));
            gap: 16px;
            margin-bottom: 120px;
        }
        .service-card {
            background: var(--card-bg);
            border: 1px solid var(--card-border);
            border-radius: 14px;
            padding: 20px;
            cursor: pointer;
            transition: all 0.3s;
            text-align: center;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            min-height: 130px;
        }
        .service-card:hover {
            border-color: var(--primary);
            transform: translateY(-4px);
            box-shadow: 0 8px 20px rgba(212, 138, 151, 0.2);
        }
        .service-name {
            font-weight: 600;
            font-size: 15px;
            line-height: 1.4;
            color: #ffffff;
            margin-bottom: 10px;
        }
        .service-price {
            font-size: 16px;
            color: var(--accent);
            font-weight: 700;
        }
        
        /* Floating Bill Bar */
        .session-drawer {
            position: fixed;
            bottom: 20px;
            left: 50%;
            transform: translateX(-50%);
            width: 90%;
            max-width: 600px;
            background: rgba(28, 13, 18, 0.95);
            backdrop-filter: blur(20px);
            -webkit-backdrop-filter: blur(20px);
            border: 1px solid var(--primary);
            border-radius: 20px;
            padding: 18px 24px;
            box-shadow: 0 15px 40px rgba(0,0,0,0.6);
            z-index: 100;
            display: flex;
            flex-direction: column;
            gap: 12px;
        }
        .drawer-top {
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 1px solid var(--card-border);
            padding-bottom: 8px;
        }
        .drawer-title {
            font-weight: 700;
            font-size: 16px;
            color: #ffffff;
        }
        .drawer-list {
            max-height: 100px;
            overflow-y: auto;
            display: flex;
            flex-direction: column;
            gap: 6px;
        }
        .drawer-item {
            font-size: 13px;
            display: flex;
            justify-content: space-between;
            color: var(--text-sub);
        }
        .drawer-actions {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-top: 5px;
        }
        .drawer-total {
            font-weight: 700;
            font-size: 18px;
            color: var(--accent);
        }
        .btn-finish {
            padding: 10px 24px;
            background: #6bb393;
            color: white;
            border: none;
            border-radius: 8px;
            font-weight: 700;
            cursor: pointer;
            transition: all 0.2s;
            box-shadow: 0 4px 10px rgba(107, 179, 147, 0.3);
        }
        .btn-finish:hover {
            background: #5aa282;
            transform: translateY(-1px);
        }
        .btn-back {
            background: transparent;
            border: 1px solid var(--card-border);
            color: var(--text-sub);
            padding: 8px 16px;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 600;
            font-size: 14px;
            transition: all 0.2s;
        }
        .btn-back:hover {
            color: #ffffff;
            border-color: #ffffff;
        }
        
        /* Modal Popup */
        .modal-overlay {
            position: fixed;
            inset: 0;
            background: rgba(0,0,0,0.7);
            backdrop-filter: blur(8px);
            -webkit-backdrop-filter: blur(8px);
            display: none;
            align-items: center;
            justify-content: center;
            z-index: 1000;
        }
        .modal {
            background: #231217;
            border: 1px solid var(--primary);
            border-radius: 20px;
            padding: 26px;
            width: 90%;
            max-width: 400px;
            box-shadow: 0 20px 50px rgba(0,0,0,0.6);
            text-align: center;
        }
        .modal-title {
            font-weight: 700;
            font-size: 18px;
            margin-bottom: 12px;
            color: #ffffff;
        }
        .modal-desc {
            font-size: 14px;
            color: var(--text-sub);
            margin-bottom: 18px;
        }
        .modal-input {
            width: 100%;
            padding: 12px;
            background: rgba(0,0,0,0.3);
            border: 1px solid var(--card-border);
            border-radius: 8px;
            color: #ffffff;
            font-size: 16px;
            outline: none;
            margin-bottom: 20px;
            text-align: center;
            font-family: inherit;
        }
        .modal-input:focus {
            border-color: var(--primary);
        }
        .modal-buttons {
            display: flex;
            gap: 12px;
        }
        .btn-modal-submit {
            flex: 1;
            padding: 12px;
            background: var(--primary);
            color: white;
            border: none;
            border-radius: 8px;
            font-weight: 700;
            cursor: pointer;
        }
        .btn-modal-cancel {
            flex: 1;
            padding: 12px;
            background: transparent;
            border: 1px solid var(--card-border);
            color: var(--text-sub);
            border-radius: 8px;
            cursor: pointer;
        }
        
        /* Face Recognition Camera View */
        .camera-wrapper {
            position: relative;
            width: 100%;
            max-width: 480px;
            margin: 0 auto 20px auto;
            background: #000;
            border-radius: 16px;
            overflow: hidden;
            border: 2px solid var(--primary);
            aspect-ratio: 4/3;
            box-shadow: 0 10px 30px rgba(0,0,0,0.5);
        }
        #webcamVideo {
            width: 100%;
            height: 100%;
            object-fit: cover;
        }
        .scan-line {
            position: absolute;
            top: 0;
            left: 0;
            width: 100%;
            height: 4px;
            background: var(--accent);
            box-shadow: 0 0 15px var(--accent);
            animation: scanAnimation 2s linear infinite;
            pointer-events: none;
        }
        @keyframes scanAnimation {
            0% { top: 0%; }
            50% { top: 100%; }
            100% { top: 0%; }
        }
        .status-info, .status-scanning, .status-checked-in, .status-checked-out, .status-error {
            border-radius: 12px;
            padding: 16px;
            text-align: center;
            font-size: 15px;
            margin-bottom: 20px;
            min-height: 60px;
            display: flex;
            align-items: center;
            justify-content: center;
            line-height: 1.5;
            font-weight: 500;
        }
        .status-info {
            background: rgba(255, 255, 255, 0.05);
            border: 1px solid var(--card-border);
        }
        .status-scanning {
            background: rgba(229, 169, 180, 0.1);
            border: 1px solid var(--accent);
        }
        .status-checked-in {
            background: rgba(107, 179, 147, 0.15);
            border: 1px solid #6bb393;
            color: #a8ebd1;
        }
        .status-checked-out {
            background: rgba(212, 138, 151, 0.15);
            border: 1px solid var(--primary);
            color: #f7cbd2;
        }
        .status-error {
            background: rgba(255, 99, 71, 0.15);
            border: 1px solid #FF6347;
            color: #ffb3a7;
        }
        
        /* Toast Notifications */
        .toast-container {
            position: fixed;
            bottom: 140px;
            left: 20px;
            z-index: 1000;
            display: flex;
            flex-direction: column;
            gap: 10px;
        }
        .toast {
            background: rgba(43, 24, 30, 0.9);
            border-left: 4px solid var(--primary);
            padding: 12px 20px;
            border-radius: 8px;
            color: #ffffff;
            box-shadow: 0 4px 15px rgba(0,0,0,0.4);
            font-weight: 500;
            animation: slideIn 0.3s forwards;
            min-width: 250px;
        }
        @keyframes slideIn {
            from { transform: translateX(-120%); opacity: 0; }
            to { transform: translateX(0); opacity: 1; }
        }
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>KAYA KALP</h1>
            <div class="subtitle">Beauty Parlour Portal</div>
        </header>

        <!-- View 1: Dashboard (Initially Hidden) -->
        <div id="dashboardView" class="view-panel">
            <div class="card">
                <div class="table-responsive">
                    <table>
                        <thead>
                            <tr>
                                <th style="width: 50px;">Select</th>
                                <th>Name of Beautician</th>
                                <th>Services Today</th>
                                <th>Services Month</th>
                                <th>Services Total</th>
                                <th>Earnings Today</th>
                                <th>Earnings Month</th>
                            </tr>
                        </thead>
                        <tbody id="beauticianTableBody">
                            <!-- Populated dynamically -->
                        </tbody>
                    </table>
                </div>
                <div style="display:flex; justify-content:center; gap: 15px; flex-wrap:wrap; margin-top:20px;">
                    <button class="btn-proceed" style="margin:0;" onclick="proceedToRateList()">Proceed to Rate List</button>
                    <button class="btn-proceed" style="margin:0; background: linear-gradient(135deg, var(--accent) 0%, #d48a97 100%); color: #12090b; box-shadow: 0 4px 12px var(--accent-glow);" onclick="openFaceAttendance()">Face Attendance Check</button>
                    <button class="btn-proceed" style="margin:0; background: #6c757d;" onclick="logoutBeautician()">Lock / Switch User</button>
                </div>
            </div>

            <!-- Bottom Stats -->
            <div class="bottom-stats">
                <div class="stat-card" style="text-align: center;">
                    <div class="stat-title" id="todayStatTitle">Total Money Made Today</div>
                    <div class="stat-value" id="overallTodayTotal">Rs. 0</div>
                </div>
                <div class="stat-card">
                    <div class="stat-title" id="monthStatTitle">Monthly Combined Earnings</div>
                    <div class="stat-list" id="monthlyEarningsList">
                        <!-- Populated dynamically -->
                    </div>
                </div>
            </div>
        </div>

        <!-- View 2: Rate List -->
        <div id="rateListView" class="view-panel">
            <div class="rate-header">
                <button class="btn-back" onclick="backToDashboard()">⬅ Back to Dashboard</button>
                <div class="beautician-badge" id="currentBeauticianBadge">Beautician: </div>
            </div>

            <div class="category-tabs" id="categoryTabs">
                <!-- Tabs generated dynamically -->
            </div>

            <div class="services-grid" id="servicesGrid">
                <!-- Services generated dynamically -->
            </div>

            <!-- Floating Cart/Session Drawer -->
            <div class="session-drawer" id="sessionDrawer" style="display: none;">
                <div class="drawer-top">
                    <div class="drawer-title">Current Customer Session</div>
                    <button class="btn-back" style="padding: 4px 8px; font-size: 11px;" onclick="clearSessionCart()">Clear</button>
                </div>
                <div class="drawer-list" id="sessionList">
                    <!-- Session items -->
                </div>
                <div class="drawer-actions">
                    <div class="drawer-total" id="sessionTotalDisplay">Total: Rs. 0</div>
                    <button class="btn-finish" onclick="finishSession()">Finish Customer</button>
                </div>
            </div>
        </div>

        <!-- View 3: Face Attendance Camera View (Default Entry Page) -->
        <div id="faceAttendanceView" class="view-panel active">
            <div class="rate-header">
                <button class="btn-back" id="faceBackBtn" style="display: none;" onclick="closeFaceAttendance()">⬅ Back to Dashboard</button>
                <div class="beautician-badge">Face Scanner Active</div>
            </div>

            <div class="card" style="text-align: center; max-width: 600px; margin: 0 auto;">
                <div class="camera-wrapper">
                    <video id="webcamVideo" autoplay playsinline muted></video>
                    <div class="scan-line"></div>
                </div>

                <div id="cameraStatus" class="status-info">
                    Point at your face and click <strong>Scan Face</strong> to check-in.
                </div>

                <button class="btn-proceed" style="max-width: 250px; margin: 10px auto;" onclick="scanFace()">Scan Face</button>
            </div>
        </div>
    </div>

    <!-- Modal for Ask Prices -->
    <div class="modal-overlay" id="priceModal">
        <div class="modal">
            <div class="modal-title" id="modalTitle">Custom Service Price</div>
            <div class="modal-desc" id="modalDesc">Enter service price in Rupees:</div>
            <input type="number" class="modal-input" id="modalPriceInput" min="0" placeholder="e.g. 500">
            <div class="modal-buttons">
                <button class="btn-modal-cancel" onclick="closePriceModal()">Cancel</button>
                <button class="btn-modal-submit" onclick="submitPriceModal()">Submit Price</button>
            </div>
        </div>
    </div>

    <!-- Toast container -->
    <div class="toast-container" id="toastContainer"></div>

    <script>
        let currentBeautician = null;
        let rateListData = null;
        let currentCategory = "";
        let sessionCart = []; 
        let sessionTotal = 0;

        let pendingService = null;
        let pendingCategory = null;

        let cameraStream = null;
        let scanning = false;
        let authenticated = false;

        document.addEventListener("DOMContentLoaded", () => {
            loadDashboard();
            loadRateListData();
            startCamera();
        });

        function loadDashboard() {
            const payload = {};
            if (currentBeautician) {
                payload.beautician_id = currentBeautician.id;
            }
            fetch("/api/get_dashboard", {
                method: "POST",
                headers: { "Content-Type": "application/json" },
                body: JSON.stringify(payload)
            })
                .then(res => res.json())
                .then(data => {
                    if (data.success) {
                        renderDashboard(data);
                    }
                })
                .catch(err => {
                    showToast("Error loading dashboard data.");
                });
        }

        function loadRateListData() {
            fetch("/api/get_rate_list")
                .then(res => res.json())
                .then(data => {
                    rateListData = data;
                });
        }

        function renderDashboard(data) {
            const tbody = document.getElementById("beauticianTableBody");
            tbody.innerHTML = "";
            
            if (data.beauticians.length === 0) {
                tbody.innerHTML = `<tr><td colspan="7" style="text-align:center;color:var(--text-sub);">No beauticians registered on the server. Please add them in the Desktop Admin GUI.</td></tr>`;
            } else {
                data.beauticians.forEach(b => {
                    const isChecked = currentBeautician && currentBeautician.id === b.id ? "checked" : "";
                    const row = document.createElement("tr");
                    row.innerHTML = `
                        <td>
                            <label class="checkbox-container">
                                <input type="checkbox" name="beauticianSelect" value="${b.id}" data-name="${b.name}" ${isChecked} onchange="handleSingleSelect(this)">
                                <span class="checkmark"></span>
                            </label>
                        </td>
                        <td style="font-weight: 600; color: #ffffff;">${b.name}</td>
                        <td>${b.stats.services_today}</td>
                        <td>${b.stats.services_month}</td>
                        <td>${b.stats.services_total}</td>
                        <td style="color:var(--accent); font-weight:600;">Rs. ${b.stats.money_today}</td>
                        <td style="color:var(--accent); font-weight:600;">Rs. ${b.stats.money_month}</td>
                    `;
                    tbody.appendChild(row);
                });
            }

            document.getElementById("overallTodayTotal").innerText = `Rs. ${data.overall.total_today}`;
            
            const list = document.getElementById("monthlyEarningsList");
            list.innerHTML = "";
            if (data.overall.monthly_breakdown.length === 0) {
                list.innerHTML = `<div style="text-align:center; color:var(--text-sub); padding:20px 0;">No transaction records found.</div>`;
            } else {
                data.overall.monthly_breakdown.forEach(item => {
                    const row = document.createElement("div");
                    row.className = "stat-item";
                    row.innerHTML = `
                        <span style="font-weight: 500;">${item.month_str}</span>
                        <span style="font-weight: 700; color: var(--accent);">Rs. ${item.amount}</span>
                    `;
                    list.appendChild(row);
                });
            }

            // Always show stats panel; relabel based on authentication state
            document.querySelector(".bottom-stats").style.display = "grid";
            if (currentBeautician) {
                document.getElementById("todayStatTitle").innerText = "Your Earnings Today";
                document.getElementById("monthStatTitle").innerText = "Your Earnings This Month";
            } else {
                document.getElementById("todayStatTitle").innerText = "Total Money Made Today";
                document.getElementById("monthStatTitle").innerText = "Monthly Combined Earnings";
            }
        }

        function handleSingleSelect(checkbox) {
            if (checkbox.checked) {
                const checkboxes = document.getElementsByName("beauticianSelect");
                checkboxes.forEach(cb => {
                    if (cb !== checkbox) cb.checked = false;
                });
            }
        }

        document.addEventListener("change", (e) => {
            if(e.target.name === "beauticianSelect") {
                handleSingleSelect(e.target);
            }
        });

        function getSelectedBeautician() {
            const checkboxes = document.getElementsByName("beauticianSelect");
            for (let cb of checkboxes) {
                if (cb.checked) {
                    return { id: cb.value, name: cb.getAttribute("data-name") };
                }
            }
            return null;
        }

        function proceedToRateList() {
            const selected = getSelectedBeautician();
            if (!selected) {
                alert("Please select a Beautician first by checking the checkbox next to her name.");
                return;
            }
            if (!rateListData) {
                showToast("Rate list data is not loaded yet. Retrying...");
                loadRateListData();
                return;
            }
            currentBeautician = selected;
            
            document.getElementById("currentBeauticianBadge").innerText = `Beautician: ${currentBeautician.name}`;
            document.getElementById("dashboardView").classList.remove("active");
            document.getElementById("rateListView").classList.add("active");
            
            resetSessionCart();
            renderCategoryTabs();
        }

        function backToDashboard() {
            if (sessionCart.length > 0) {
                if (!confirm("You have service logs on screen. Going back to dashboard will discard these items. Do you want to go back?")) {
                    return;
                }
            }
            document.getElementById("rateListView").classList.remove("active");
            document.getElementById("dashboardView").classList.add("active");
            loadDashboard();
            currentBeautician = null;
            resetSessionCart();
        }

        function renderCategoryTabs() {
            const container = document.getElementById("categoryTabs");
            container.innerHTML = "";
            const categories = Object.keys(rateListData);
            
            if (categories.length > 0) {
                currentCategory = categories[0];
                categories.forEach(cat => {
                    const tab = document.createElement("div");
                    tab.className = `category-tab ${cat === currentCategory ? "active" : ""}`;
                    tab.innerText = cat;
                    tab.onclick = () => {
                        document.querySelectorAll(".category-tab").forEach(t => t.classList.remove("active"));
                        tab.classList.add("active");
                        currentCategory = cat;
                        renderServices();
                    };
                    container.appendChild(tab);
                });
                renderServices();
            }
        }

        function renderServices() {
            const grid = document.getElementById("servicesGrid");
            grid.innerHTML = "";
            
            const services = rateListData[currentCategory];
            services.forEach(serv => {
                const card = document.createElement("div");
                card.className = "service-card";
                card.onclick = () => handleServiceClick(serv);
                
                const rateText = typeof serv.rate === 'number' ? `Rs. ${serv.rate}` : "Ask Price";
                card.innerHTML = `
                    <div class="service-name">${serv.name}</div>
                    <div class="service-price">${rateText}</div>
                `;
                grid.appendChild(card);
            });
        }

        function handleServiceClick(service) {
            if (service.rate === "Ask") {
                openPriceModal(service, currentCategory);
            } else {
                addLocalService(currentCategory, service.name, service.rate);
            }
        }

        function openPriceModal(service, category) {
            pendingService = service;
            pendingCategory = category;
            document.getElementById("modalTitle").innerText = service.name;
            document.getElementById("modalPriceInput").value = "";
            document.getElementById("priceModal").style.display = "flex";
            document.getElementById("modalPriceInput").focus();
        }

        function closePriceModal() {
            document.getElementById("priceModal").style.display = "none";
            pendingService = null;
            pendingCategory = null;
        }

        function submitPriceModal() {
            const input = document.getElementById("modalPriceInput");
            const price = parseFloat(input.value);
            if (isNaN(price) || price < 0) {
                alert("Please enter a valid positive number for the service price.");
                return;
            }
            const servName = pendingService.name;
            const cat = pendingCategory;
            
            closePriceModal();
            addLocalService(cat, servName, price);
        }

        function addLocalService(category, name, rate) {
            sessionCart.push({ category: category, name: name, rate: rate });
            sessionTotal += rate;
            
            showToast(`Added: ${name} - Rs. ${rate}`);
            document.getElementById("sessionDrawer").style.display = "flex";
            
            const list = document.getElementById("sessionList");
            const item = document.createElement("div");
            item.className = "drawer-item";
            item.innerHTML = `
                <span>${name}</span>
                <span style="font-weight:600;">Rs. ${rate}</span>
            `;
            list.appendChild(item);
            list.scrollTop = list.scrollHeight;
            
            document.getElementById("sessionTotalDisplay").innerText = `Total: Rs. ${sessionTotal}`;
        }

        function resetSessionCart() {
            sessionCart = [];
            sessionTotal = 0;
            document.getElementById("sessionList").innerHTML = "";
            document.getElementById("sessionDrawer").style.display = "none";
        }

        function clearSessionCart() {
            if (sessionCart.length === 0) return;
            const removed = sessionCart.pop();
            sessionTotal = Math.max(0, sessionTotal - removed.rate);
            const list = document.getElementById("sessionList");
            if (list.lastElementChild) list.lastElementChild.remove();
            document.getElementById("sessionTotalDisplay").innerText = `Total: Rs. ${sessionTotal}`;
            if (sessionCart.length === 0) {
                document.getElementById("sessionDrawer").style.display = "none";
            }
        }

        function finishSession() {
            if (sessionCart.length === 0) {
                backToDashboard();
                return;
            }
            
            fetch("/api/log_session", {
                method: "POST",
                headers: { "Content-Type": "application/json" },
                body: JSON.stringify({
                    beautician_id: currentBeautician.id,
                    beautician_name: currentBeautician.name,
                    services: sessionCart
                })
            })
            .then(res => res.json())
            .then(data => {
                if (data.success) {
                    alert(`Customer bill of Rs. ${sessionTotal} logged successfully! Returning to dashboard.`);
                    backToDashboard();
                } else {
                    alert("Failed to save session: " + (data.error || "Server error"));
                }
            })
            .catch(() => {
                alert("Failed to connect to server to log session.");
            });
        }

        function showToast(message) {
            const container = document.getElementById("toastContainer");
            const toast = document.createElement("div");
            toast.className = "toast";
            toast.innerText = message;
            container.appendChild(toast);
            
            setTimeout(() => {
                toast.style.animation = "slideIn 0.3s reverse forwards";
                setTimeout(() => {
                    toast.remove();
                }, 300);
            }, 3000);
        }

        // ====== Face Attendance Functions ======
        async function startCamera() {
            const video = document.getElementById("webcamVideo");
            try {
                cameraStream = await navigator.mediaDevices.getUserMedia({
                    video: { facingMode: "user", width: { ideal: 640 }, height: { ideal: 480 } },
                    audio: false
                });
                video.srcObject = cameraStream;
                document.getElementById("cameraStatus").innerHTML = "Point at your face and click <strong>Scan Face</strong> to check-in or check-out.";
                document.getElementById("cameraStatus").className = "status-info";
            } catch (e) {
                document.getElementById("cameraStatus").innerHTML = "Webcam error: " + e.message;
                document.getElementById("cameraStatus").className = "status-error";
            }
        }

        function stopCamera() {
            if (cameraStream) {
                cameraStream.getTracks().forEach(track => track.stop());
                cameraStream = null;
            }
            const video = document.getElementById("webcamVideo");
            if (video) video.srcObject = null;
        }

        function openFaceAttendance() {
            document.getElementById("dashboardView").classList.remove("active");
            document.getElementById("faceAttendanceView").classList.add("active");
            startCamera();
        }

        function closeFaceAttendance() {
            stopCamera();
            document.getElementById("faceAttendanceView").classList.remove("active");
            document.getElementById("dashboardView").classList.add("active");
            loadDashboard(); 
        }

        function logoutBeautician() {
            authenticated = false;
            currentBeautician = null;
            document.getElementById("dashboardView").classList.remove("active");
            document.getElementById("faceAttendanceView").classList.add("active");
            document.getElementById("faceBackBtn").style.display = "none";
            startCamera();
        }

        async function scanFace() {
            if (scanning) return;
            const video = document.getElementById("webcamVideo");
            if (!video.srcObject) {
                showToast("Webcam is not started.");
                return;
            }
            
            scanning = true;
            const canvas = document.createElement("canvas");
            canvas.width = video.videoWidth || 640;
            canvas.height = video.videoHeight || 480;
            const ctx = canvas.getContext("2d");
            ctx.drawImage(video, 0, 0, canvas.width, canvas.height);
            
            const imageData = canvas.toDataURL("image/jpeg", 0.85);
            
            const statusBox = document.getElementById("cameraStatus");
            statusBox.innerHTML = "🔍 Scanning face...";
            statusBox.className = "status-scanning";
            
            fetch("/api/recognize", {
                method: "POST",
                headers: { "Content-Type": "application/json" },
                body: JSON.stringify({ image: imageData })
            })
            .then(res => res.json())
            .then(data => {
                scanning = false;
                if (data.success) {
                    const timeStr = data.time;
                    const action = data.action; 
                    const name = data.name;
                    
                    let greeting = "";
                    let colorClass = "";
                    if (action === "Checked In") {
                        greeting = `🌸 Welcome, <strong>${name}</strong>!<br>Checked in successfully at ${timeStr}.`;
                        colorClass = "status-checked-in";
                        
                        authenticated = true;
                        currentBeautician = { id: data.id, name: name };
                        
                        setTimeout(() => {
                            document.getElementById("faceAttendanceView").classList.remove("active");
                            document.getElementById("dashboardView").classList.add("active");
                            document.getElementById("faceBackBtn").style.display = "block";
                            loadDashboard();
                            stopCamera();
                        }, 2000);
                    } else {
                        greeting = `👋 Goodbye, <strong>${name}</strong>!<br>Checked out successfully at ${timeStr}.`;
                        colorClass = "status-checked-out";
                        
                        authenticated = false;
                        currentBeautician = null;
                        document.getElementById("faceBackBtn").style.display = "none";
                    }
                    statusBox.innerHTML = greeting;
                    statusBox.className = colorClass;
                    showToast(`${name}: ${action}`);
                } else {
                    statusBox.innerHTML = "⚠ " + (data.error || "Scanning failed.");
                    statusBox.className = "status-error";
                }
            })
            .catch(err => {
                scanning = false;
                statusBox.innerHTML = "⚠ Connection error.";
                statusBox.className = "status-error";
            });
        }
    </script>
</body>
</html>"""

# ----------- Tkinter Desktop GUI Classes -----------
class FaceRecognitionWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Face Recognition Attendance Scanner")
        self.geometry("640x520")
        self.resizable(False, False)
        self.configure(bg="#fdf4f5")
        
        # Reload known faces first
        load_known_faces()
        
        # Display label
        self.video_label = tk.Label(self, bg="black")
        self.video_label.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.status_label = tk.Label(self, text="Initializing Camera...", bg="#fdf4f5", font=("Outfit", 10, "bold"))
        self.status_label.pack(pady=5)
        
        self.cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
        if not self.cap.isOpened():
            self.cap = cv2.VideoCapture(0)
            
        if not self.cap.isOpened():
            messagebox.showerror("Camera Error", "Could not open local camera. Please verify webcam is connected.")
            self.destroy()
            return
            
        self.running = True
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.update_frame()
        
    def update_frame(self):
        if not self.running:
            return
            
        ret, frame = self.cap.read()
        if not ret:
            self.after(15, self.update_frame)
            return
            
        small_frame = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)
        rgb_small = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        
        face_locations = face_recognition.face_locations(rgb_small)
        
        if face_locations:
            face_encodings = face_recognition.face_encodings(rgb_small, face_locations)
            if face_encodings and known_encodings:
                enc = face_encodings[0]
                distances = face_recognition.face_distance(known_encodings, enc)
                best_idx = np.argmin(distances) if len(distances) > 0 else None
                
                if best_idx is not None and distances[best_idx] < TOLERANCE:
                    matched_id = known_ids[best_idx]
                    matched_name = known_names.get(matched_id, "Beautician")
                    
                    top, right, bottom, left = face_locations[0]
                    top *= 2
                    right *= 2
                    bottom *= 2
                    left *= 2
                    cv2.rectangle(frame, (left, top), (right, bottom), (212, 138, 151), 2)
                    cv2.putText(frame, matched_name, (left, top - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (212, 138, 151), 2)
                    
                    success, action, time_logged = log_beautician_attendance_face(matched_id, matched_name)
                    if success:
                        self.running = False
                        self.cap.release()
                        self.destroy()
                        
                        # Show Tkinter messagebox
                        messagebox.showinfo("Attendance Logged", f"Attendance logged successfully for {matched_name} ({action} at {time_logged}).")
                        return
                    else:
                        self.running = False
                        self.cap.release()
                        self.destroy()
                        
                        # Show error messagebox
                        messagebox.showerror("Attendance Error", f"Failed to log attendance for {matched_name}:\n{action}")
                        return
                    
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(rgb_frame)
        img = img.resize((620, 440))
        img_tk = ImageTk.PhotoImage(img)
        self.video_label.config(image=img_tk)
        self.video_label.image = img_tk
        
        self.status_label.config(text="Align face in camera view to check-in/out.")
        self.after(15, self.update_frame)
        
    def on_close(self):
        self.running = False
        if self.cap.isOpened():
            self.cap.release()
        self.destroy()


class RateListWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Manage Rate List")
        self.geometry("800x600")
        self.configure(bg="#fdf4f5")
        self.transient(master)
        self.grab_set()

        # Local copy of rate list
        self.local_rates = json.loads(json.dumps(RATE_LIST))

        # Main horizontal layout pane
        main_pane = tk.PanedWindow(self, orient=tk.HORIZONTAL, bg="#fdf4f5", bd=0)
        main_pane.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # --- Left Pane (Categories) ---
        left_frame = ttk.LabelFrame(main_pane, text="Categories")
        main_pane.add(left_frame, minsize=220)

        self.cat_listbox = tk.Listbox(left_frame, font=("Outfit", 10), bg="white", selectbackground="#e5a9b4", selectforeground="black")
        self.cat_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.cat_listbox.bind("<<ListboxSelect>>", self.on_category_select)

        cat_btn_frame = ttk.Frame(left_frame)
        cat_btn_frame.pack(fill=tk.X, padx=5, pady=5)
        tk.Button(cat_btn_frame, text="Add Category", command=self.add_category, bg="#d48a97", fg="white", font=("Outfit", 9, "bold")).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
        tk.Button(cat_btn_frame, text="Delete Category", command=self.delete_category, bg="#F08080", fg="white", font=("Outfit", 9, "bold")).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        # --- Right Pane (Services) ---
        right_frame = ttk.LabelFrame(main_pane, text="Services & Rates")
        main_pane.add(right_frame, minsize=450)

        table_frame = ttk.Frame(right_frame)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.serv_table = ttk.Treeview(table_frame, columns=["Name", "Rate"], show="headings")
        self.serv_table.heading("Name", text="Service Name")
        self.serv_table.heading("Rate", text="Rate (or 'Ask')")
        self.serv_table.column("Name", width=300)
        self.serv_table.column("Rate", width=120)
        self.serv_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.serv_table.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.serv_table.configure(yscrollcommand=vsb.set)
        self.serv_table.bind("<Double-1>", self.edit_service)

        serv_btn_frame = ttk.Frame(right_frame)
        serv_btn_frame.pack(fill=tk.X, padx=5, pady=5)
        tk.Button(serv_btn_frame, text="Add Service", command=self.add_service, bg="#d48a97", fg="white", font=("Outfit", 9, "bold")).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
        tk.Button(serv_btn_frame, text="Edit Selected", command=self.edit_service, bg="#d48a97", fg="white", font=("Outfit", 9, "bold")).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
        tk.Button(serv_btn_frame, text="Delete Service", command=self.delete_service, bg="#F08080", fg="white", font=("Outfit", 9, "bold")).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        # --- Bottom Panel (Save & Close) ---
        bottom_frame = ttk.Frame(self)
        bottom_frame.pack(fill=tk.X, padx=10, pady=10)
        tk.Button(bottom_frame, text="Save Rate List Changes", command=self.save_changes, bg="#d48a97", fg="white", font=("Outfit", 10, "bold")).pack(side=tk.RIGHT, padx=5)
        tk.Button(bottom_frame, text="Cancel", command=self.destroy, bg="#ccc", fg="black", font=("Outfit", 10)).pack(side=tk.RIGHT, padx=5)

        self.populate_categories()

    def populate_categories(self):
        self.cat_listbox.delete(0, tk.END)
        for cat in sorted(self.local_rates.keys()):
            self.cat_listbox.insert(tk.END, cat)
        if self.cat_listbox.size() > 0:
            self.cat_listbox.selection_set(0)
            self.on_category_select()

    def on_category_select(self, event=None):
        selected = self.cat_listbox.curselection()
        # Clear services treeview
        self.serv_table.delete(*self.serv_table.get_children())
        if not selected:
            return
        cat = self.cat_listbox.get(selected[0])
        services = self.local_rates.get(cat, [])
        for idx, s in enumerate(services):
            self.serv_table.insert("", tk.END, iid=str(idx), values=[s["name"], s["rate"]])

    def add_category(self):
        from tkinter import simpledialog
        cat = simpledialog.askstring("Add Category", "Enter new category name:")
        if not cat:
            return
        cat = cat.strip()
        if cat in self.local_rates:
            messagebox.showerror("Error", "Category already exists.")
            return
        self.local_rates[cat] = []
        self.populate_categories()
        # Find index of the newly added category and select it
        for idx in range(self.cat_listbox.size()):
            if self.cat_listbox.get(idx) == cat:
                self.cat_listbox.selection_clear(0, tk.END)
                self.cat_listbox.selection_set(idx)
                self.on_category_select()
                break

    def delete_category(self):
        selected = self.cat_listbox.curselection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a category to delete.")
            return
        cat = self.cat_listbox.get(selected[0])
        if not messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete category '{cat}'? All services in it will be lost."):
            return
        del self.local_rates[cat]
        self.populate_categories()

    def add_service(self):
        from tkinter import simpledialog
        selected = self.cat_listbox.curselection()
        if not selected:
            messagebox.showwarning("Warning", "Select a category first.")
            return
        cat = self.cat_listbox.get(selected[0])
        
        name = simpledialog.askstring("Add Service", "Enter service name:")
        if not name:
            return
        name = name.strip()
        
        rate_str = simpledialog.askstring("Add Service", "Enter service rate (or 'Ask'):")
        if rate_str is None:
            return
        rate_str = rate_str.strip()
        
        if rate_str.lower() == "ask":
            rate = "Ask"
        else:
            try:
                rate = int(rate_str)
            except ValueError:
                try:
                    rate = float(rate_str)
                except ValueError:
                    rate = "Ask"
                    
        self.local_rates[cat].append({"name": name, "rate": rate})
        self.on_category_select()

    def edit_service(self, event=None):
        from tkinter import simpledialog
        selected_cat = self.cat_listbox.curselection()
        if not selected_cat:
            return
        cat = self.cat_listbox.get(selected_cat[0])
        
        selected_serv = self.serv_table.selection()
        if not selected_serv:
            if event is None:
                messagebox.showwarning("Warning", "Select a service to edit.")
            return
            
        idx = int(selected_serv[0])
        current_service = self.local_rates[cat][idx]
        
        name = simpledialog.askstring("Edit Service", "Service name:", initialvalue=current_service["name"])
        if not name:
            return
        name = name.strip()
        
        rate_str = simpledialog.askstring("Edit Service", "Rate (or 'Ask'):", initialvalue=str(current_service["rate"]))
        if rate_str is None:
            return
        rate_str = rate_str.strip()
        
        if rate_str.lower() == "ask":
            rate = "Ask"
        else:
            try:
                rate = int(rate_str)
            except ValueError:
                try:
                    rate = float(rate_str)
                except ValueError:
                    rate = "Ask"
                    
        self.local_rates[cat][idx] = {"name": name, "rate": rate}
        self.on_category_select()

    def delete_service(self):
        selected_cat = self.cat_listbox.curselection()
        if not selected_cat:
            return
        cat = self.cat_listbox.get(selected_cat[0])
        
        selected_serv = self.serv_table.selection()
        if not selected_serv:
            messagebox.showwarning("Warning", "Select a service to delete.")
            return
            
        idx = int(selected_serv[0])
        service_name = self.local_rates[cat][idx]["name"]
        if not messagebox.askyesno("Confirm Delete", f"Delete service '{service_name}'?"):
            return
            
        self.local_rates[cat].pop(idx)
        self.on_category_select()

    def save_changes(self):
        global RATE_LIST
        RATE_LIST.clear()
        RATE_LIST.update(self.local_rates)
        save_rates()
        messagebox.showinfo("Success", "Rate list changes saved successfully!")
        self.destroy()


class AttendanceWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Attendance & Service Records")
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{int(sw*0.75)}x{int(sh*0.7)}")
        self.configure(bg="#fdf4f5")
        
        style = ttk.Style()
        style.configure("Parlour.TNotebook", background="#fdf4f5")
        style.configure("Parlour.TNotebook.Tab", background="#e5a9b4", foreground="black", padding=[12, 6], font=("Outfit", 10, "bold"))
        
        notebook = ttk.Notebook(self, style="Parlour.TNotebook")
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        attendance_frame = ttk.Frame(notebook)
        notebook.add(attendance_frame, text="Daily Attendance")
        self.build_attendance_tab(attendance_frame)
        
        services_frame = ttk.Frame(notebook)
        notebook.add(services_frame, text="Service Transactions")
        self.build_services_tab(services_frame)

    def build_attendance_tab(self, frame):
        search_frame = ttk.Frame(frame)
        search_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT)
        self.att_search_var = tk.StringVar()
        self.att_search_entry = ttk.Entry(search_frame, textvariable=self.att_search_var, width=30)
        self.att_search_entry.pack(side=tk.LEFT, padx=5)
        self.att_search_var.trace_add("write", self.filter_attendance)
        
        table_frame = ttk.Frame(frame)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        self.att_table = ttk.Treeview(table_frame, columns=["S.no", "Date", "Beautician ID", "Name", "Check-in Time", "Check-out Time"], show="headings")
        self.att_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.att_table.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.att_table.configure(yscrollcommand=vsb.set)
        
        for col in ["S.no", "Date", "Beautician ID", "Name", "Check-in Time", "Check-out Time"]:
            self.att_table.heading(col, text=col)
            self.att_table.column(col, width=120, anchor=tk.W)
            
        self.load_attendance_data()

    def build_services_tab(self, frame):
        search_frame = ttk.Frame(frame)
        search_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT)
        self.serv_search_var = tk.StringVar()
        self.serv_search_entry = ttk.Entry(search_frame, textvariable=self.serv_search_var, width=30)
        self.serv_search_entry.pack(side=tk.LEFT, padx=5)
        self.serv_search_var.trace_add("write", self.filter_services)
        
        table_frame = ttk.Frame(frame)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        self.serv_table = ttk.Treeview(table_frame, columns=["S.no", "Date", "Time", "Beautician ID", "Name", "Category", "Service Name", "Rate"], show="headings")
        self.serv_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.serv_table.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.serv_table.configure(yscrollcommand=vsb.set)
        
        for col in ["S.no", "Date", "Time", "Beautician ID", "Name", "Category", "Service Name", "Rate"]:
            self.serv_table.heading(col, text=col)
            self.serv_table.column(col, width=100, anchor=tk.W)
            
        self.load_services_data()

    def load_attendance_data(self):
        self.att_table.delete(*self.att_table.get_children())
        self.att_data = []
        ensure_attendance_workbook()
        if os.path.exists(ATTENDANCE_FILE):
            try:
                wb = load_workbook(ATTENDANCE_FILE, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and any(row):
                        self.att_data.append(row)
                wb.close()
            except Exception as e:
                print("Error loading attendance:", e)
        self.att_data.reverse()
        self.filter_attendance()

    def filter_attendance(self, *args):
        self.att_table.delete(*self.att_table.get_children())
        search_text = self.att_search_var.get().lower()
        for row in self.att_data:
            if not search_text or any(search_text in str(cell).lower() for cell in row if cell is not None):
                self.att_table.insert("", tk.END, values=[str(c or "") for c in row])

    def load_services_data(self):
        self.serv_table.delete(*self.serv_table.get_children())
        self.serv_data = []
        ensure_services_log_workbook()
        if os.path.exists(SERVICES_LOG_FILE):
            try:
                wb = load_workbook(SERVICES_LOG_FILE, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and any(row):
                        self.serv_data.append(row)
                wb.close()
            except Exception as e:
                print("Error loading services log:", e)
        self.serv_data.reverse()
        self.filter_services()

    def filter_services(self, *args):
        self.serv_table.delete(*self.serv_table.get_children())
        search_text = self.serv_search_var.get().lower()
        for row in self.serv_data:
            if not search_text or any(search_text in str(cell).lower() for cell in row if cell is not None):
                self.serv_table.insert("", tk.END, values=[str(c or "") for c in row])


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Kaya Kalp Admin Dashboard")
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "favicon.ico")
        if os.path.exists(icon_path):
            self.iconbitmap(icon_path)
            
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{int(sw*0.75)}x{int(sh*0.65)}")
        self.configure(bg="#fdf4f5")
        
        ensure_workbook()
        ensure_attendance_workbook()
        ensure_services_log_workbook()
        
        self.inputs = {}
        self.server_url_var = tk.StringVar(value="Web URL: Initializing...")
        self.total_str_var = tk.StringVar(value="Registered Beauticians: 0")
        
        # Overall financial summary variables
        self.overall_today_var = tk.StringVar(value="Total Money Made Today: Rs. 0")
        self.overall_monthly_var = tk.StringVar(value="Combined Monthly Earnings: No records")

        header_frame = tk.Frame(self, bg="#fdf4f5")
        header_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=(6, 0))
        
        base_frame = ttk.LabelFrame(self, text="Beautician Details")
        base_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=8)
        
        ttk.Label(base_frame, text="Beautician ID").grid(row=0, column=0, padx=6, pady=4, sticky="w")
        ent_id = ttk.Entry(base_frame, width=20)
        ent_id.grid(row=1, column=0, padx=6, pady=4, sticky="we")
        self.inputs["ID"] = ent_id
        
        ttk.Label(base_frame, text="Name").grid(row=0, column=1, padx=6, pady=4, sticky="w")
        ent_name = ttk.Entry(base_frame, width=30)
        ent_name.grid(row=1, column=1, padx=6, pady=4, sticky="we")
        self.inputs["Name"] = ent_name

        btn_frame = ttk.Frame(self)
        btn_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=4)
        
        tk.Button(btn_frame, text="Save Beautician", command=self.add_data, bg="#d48a97", fg="white", font=("Outfit", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Add Photo", command=self.add_photo, bg="#d48a97", fg="white", font=("Outfit", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Face Recognition", command=self.open_face_recognition, bg="#90EE90", fg="black", font=("Outfit", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Delete Beautician", command=self.delete_data, bg="#d48a97", fg="white", font=("Outfit", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Attendance Logs", command=self.open_attendance_window, bg="#d48a97", fg="white", font=("Outfit", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Manage Rate List", command=self.open_rate_list_editor, bg="#d48a97", fg="white", font=("Outfit", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="View QR Code", command=self.show_server_qr, bg="#e5a9b4", fg="black", font=("Outfit", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Reset Database", command=self.reset_database, bg="#F08080", fg="white", font=("Outfit", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Exit", command=self.exit_app, bg="#F08080", fg="white", font=("Outfit", 9, "bold")).pack(side=tk.RIGHT, padx=5)
        
        web_info_frame = ttk.Frame(self)
        web_info_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=6)
        
        ttk.Label(web_info_frame, textvariable=self.total_str_var, font=("Outfit", 10, "bold")).pack(side=tk.LEFT, padx=5)
        ttk.Label(web_info_frame, textvariable=self.server_url_var, foreground="purple", font=("Outfit", 10, "italic")).pack(side=tk.RIGHT, padx=5)

        # Bottom summary frame
        overall_frame = ttk.LabelFrame(self, text="Overall Financial Summary")
        overall_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=(4, 10))
        
        ttk.Label(overall_frame, textvariable=self.overall_today_var, font=("Outfit", 11, "bold"), foreground="green").pack(side=tk.LEFT, padx=15, pady=8)
        ttk.Label(overall_frame, textvariable=self.overall_monthly_var, font=("Outfit", 10, "bold"), foreground="purple").pack(side=tk.RIGHT, padx=15, pady=8)

        preview_frame = ttk.LabelFrame(self, text="Registered Beauticians List")
        preview_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=8)
        
        # Left container for list
        list_container = ttk.Frame(preview_frame)
        list_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        cols = ["ID", "Name", "Services Today", "Services Month", "Services Total", "Earnings Today", "Earnings Month"]
        self.tree = ttk.Treeview(list_container, columns=cols, show="headings")
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb = ttk.Scrollbar(list_container, orient="vertical", command=self.tree.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=vsb.set)
        
        # Right container for photo preview
        self.photo_frame = ttk.LabelFrame(preview_frame, text="Beautician Photo")
        self.photo_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0), pady=0)
        
        self.photo_label = tk.Label(self.photo_frame, text="Select a beautician\nto see photo", bg="#fdf4f5", font=("Outfit", 10), width=24)
        self.photo_label.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        for c in cols:
            self.tree.heading(c, text=c)
            if c in ["ID", "Name"]:
                self.tree.column(c, width=100, anchor=tk.W)
            else:
                self.tree.column(c, width=110, anchor=tk.CENTER)
            
        self.load_preview()
        self.tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        
        # Bind double clicks for fullscreen toggle
        self.tree.bind("<Double-1>", self.open_fullscreen_photo)
        self.photo_label.bind("<Double-1>", self.open_fullscreen_photo)
        
        self.start_lan_server_thread()
        self.periodic_refresh()

    def periodic_refresh(self):
        self.load_preview()
        self.after(5000, self.periodic_refresh)

    def start_lan_server_thread(self):
        t = threading.Thread(target=self._run_server, daemon=True)
        t.start()

    def _run_server(self):
        cert_file = os.path.join(DATA_FOLDER, "server.crt")
        key_file  = os.path.join(DATA_FOLDER, "server.key")
        try:
            ensure_ssl_cert(cert_file, key_file)
        except Exception as e:
            print(f"Cannot start HTTPS server — SSL cert generation failed: {e}")
            return

        context = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
        context.load_cert_chain(cert_file, key_file)

        port = 8443
        while port < 8460:
            try:
                server = ThreadingHTTPServer(('', port), BeautyHandler)
                server.socket = context.wrap_socket(server.socket, server_side=True)
                local_ip = get_local_ip()
                url = f"https://{local_ip}:{port}"
                print(f"HTTPS Server started successfully at {url}")
                self.server_url_var.set(f"Web URL: {url}")
                try:
                    qr = qrcode.QRCode(version=1, box_size=8, border=3)
                    qr.add_data(url)
                    qr.make(fit=True)
                    img = qr.make_image(fill_color="black", back_color="white")
                    img.save(os.path.join(DATA_FOLDER, "server_qr.png"))
                except Exception as qre:
                    print(f"Auto-generation of QR code failed: {qre}")
                server.serve_forever()
                break
            except Exception as e:
                print(f"Port {port} busy, trying next:", e)
                port += 1

    def show_server_qr(self):
        url_text = self.server_url_var.get()
        if "http" not in url_text:
            messagebox.showerror("Error", "Server URL is not set.")
            return
        actual_url = url_text.split("Web URL: ")[-1].strip()
        qr_path = os.path.join(DATA_FOLDER, "server_qr.png")
        if not os.path.exists(qr_path):
            messagebox.showerror("Error", "QR code image was not generated.")
            return
            
        top = tk.Toplevel(self)
        top.title("Dashboard Web Access QR Code")
        top.geometry("350x430")
        top.configure(bg="#fdf4f5")
        top.resizable(False, False)
        top.transient(self)
        top.grab_set()
        
        lbl_title = tk.Label(top, text="Scan to Access Parlour Portal", bg="#fdf4f5", font=("Outfit", 12, "bold"))
        lbl_title.pack(pady=(15, 5))
        
        lbl_url = tk.Label(top, text=actual_url, fg="purple", bg="#fdf4f5", font=("Outfit", 10, "italic"), wraplength=320)
        lbl_url.pack(pady=(0, 15))
        
        img = Image.open(qr_path)
        photo = ImageTk.PhotoImage(img)
        lbl_qr = tk.Label(top, image=photo, bg="#fdf4f5")
        lbl_qr.image = photo
        lbl_qr.pack()
        
        lbl_note = tk.Label(top, text="Ensure your mobile device is connected to the same WiFi/LAN.", bg="#fdf4f5", font=("Outfit", 8), wraplength=300)
        lbl_note.pack(pady=(10, 10))
        
        btn_close = tk.Button(top, text="Close", command=top.destroy, bg="#F08080", fg="white", font=("Outfit", 9, "bold"), bd=0, padx=10, pady=5)
        btn_close.pack(pady=(0, 10))

    def load_preview(self):
        self.tree.delete(*self.tree.get_children())
        ensure_workbook()
        total = 0
        try:
            wb = load_workbook(FILE_NAME)
            ws = wb[SHEET_NAME]
            for r in range(2, ws.max_row + 1):
                id_val = ws.cell(row=r, column=1).value
                name_val = ws.cell(row=r, column=2).value
                if id_val:
                    b_id = str(id_val).strip()
                    name = str(name_val or "").strip()
                    stats = compute_beautician_stats(b_id)
                    row_vals = [
                        b_id,
                        name,
                        stats["services_today"],
                        stats["services_month"],
                        stats["services_total"],
                        f"Rs. {stats['money_today']}",
                        f"Rs. {stats['money_month']}"
                    ]
                    self.tree.insert("", tk.END, values=row_vals)
                    total += 1
            wb.close()
        except Exception as e:
            print("Error loading preview:", e)
        self.total_str_var.set(f"Registered Beauticians: {total}")

        # Update overall financial stats at the bottom
        try:
            stats = compute_overall_stats()
            self.overall_today_var.set(f"Total Money Made Today: Rs. {stats['total_today']}")
            breakdowns = []
            for item in stats['monthly_breakdown']:
                breakdowns.append(f"{item['month_str']}: Rs. {item['amount']}")
            if breakdowns:
                self.overall_monthly_var.set(" | ".join(breakdowns))
            else:
                self.overall_monthly_var.set("Combined Monthly Earnings: No records")
        except Exception as e:
            print("Error loading overall stats in load_preview:", e)

    def on_tree_select(self, event=None):
        selected = self.tree.selection()
        if not selected:
            return
        values = self.tree.item(selected[0], "values")
        if not values:
            return
        self.inputs["ID"].delete(0, tk.END)
        self.inputs["ID"].insert(0, values[0])
        self.inputs["Name"].delete(0, tk.END)
        self.inputs["Name"].insert(0, values[1])
        
        self.reveal_photo(values[0])

    def reveal_photo(self, b_id):
        self.photo_label.config(image="", text="No photo uploaded")
        self.photo_label.image = None
        self.current_photo_path = None
        
        valid_extensions = [".jpg", ".jpeg", ".png"]
        for ext in valid_extensions:
            img_path = os.path.join(KNOWN_FACES_DIR, f"{b_id}{ext}")
            if os.path.exists(img_path):
                try:
                    img = Image.open(img_path)
                    img.thumbnail((180, 180))
                    photo = ImageTk.PhotoImage(img)
                    self.photo_label.config(image=photo, text="")
                    self.photo_label.image = photo
                    self.current_photo_path = img_path
                    break
                except Exception as e:
                    print("Error displaying thumbnail:", e)

    def open_fullscreen_photo(self, event=None):
        if not hasattr(self, "current_photo_path") or not self.current_photo_path or not os.path.exists(self.current_photo_path):
            return
            
        fs_win = tk.Toplevel(self)
        fs_win.title("Full Screen Photo")
        fs_win.attributes("-fullscreen", True)
        fs_win.configure(bg="black")
        
        lbl = tk.Label(fs_win, bg="black")
        lbl.pack(fill=tk.BOTH, expand=True)
        
        try:
            img = Image.open(self.current_photo_path)
            screen_width = fs_win.winfo_screenwidth()
            screen_height = fs_win.winfo_screenheight()
            
            img_width, img_height = img.size
            ratio = min(screen_width / img_width, screen_height / img_height)
            new_width = int(img_width * ratio)
            new_height = int(img_height * ratio)
            
            img = img.resize((new_width, new_height), Image.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            lbl.config(image=photo)
            lbl.image = photo
        except Exception as e:
            print("Error displaying fullscreen photo:", e)
            fs_win.destroy()
            return
            
        fs_win.bind("<Double-1>", lambda e: fs_win.destroy())
        fs_win.bind("<Escape>", lambda e: fs_win.destroy())
        lbl.bind("<Double-1>", lambda e: fs_win.destroy())

    def add_data(self):
        b_id = self.inputs["ID"].get().strip()
        name = self.inputs["Name"].get().strip()
        if not b_id or not name:
            messagebox.showerror("Error", "ID and Name are required.")
            return
        
        try:
            ensure_workbook()
            wb = load_workbook(FILE_NAME)
            ws = wb[SHEET_NAME]
            
            r = find_row_by_army_no(ws, b_id)
            if r is None:
                r = ws.max_row + 1
                ws.cell(row=r, column=1, value=b_id)
            ws.cell(row=r, column=2, value=name)
            
            wb.save(FILE_NAME)
            wb.close()
            
            messagebox.showinfo("Success", f"Beautician {name} saved successfully.")
            self.inputs["ID"].delete(0, tk.END)
            self.inputs["Name"].delete(0, tk.END)
            self.load_preview()
            
            threading.Thread(target=load_known_faces, daemon=True).start()
        except Exception as e:
            messagebox.showerror("Error", f"Could not save beautician: {e}")

    def add_photo(self):
        b_id = self.inputs["ID"].get().strip()
        if not b_id:
            messagebox.showerror("Error", "Please enter/select a Beautician ID before adding a photo.")
            return
            
        file_path = filedialog.askopenfilename(
            title="Select Face Photo",
            filetypes=[("Image Files", "*.jpg *.jpeg *.png")]
        )
        if not file_path:
            return
            
        if not os.path.exists(KNOWN_FACES_DIR):
            os.makedirs(KNOWN_FACES_DIR)
            
        ext = os.path.splitext(file_path)[1].lower()
        dest_path = os.path.join(KNOWN_FACES_DIR, f"{b_id}{ext}")
        
        for e in [".jpg", ".jpeg", ".png"]:
            p = os.path.join(KNOWN_FACES_DIR, f"{b_id}{e}")
            if os.path.exists(p):
                try:
                    os.remove(p)
                except:
                    pass
                    
        try:
            shutil.copy2(file_path, dest_path)
            messagebox.showinfo("Success", f"Face photo uploaded for Beautician ID: {b_id}.")
            threading.Thread(target=load_known_faces, daemon=True).start()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save photo: {e}")

    def delete_data(self):
        b_id = self.inputs["ID"].get().strip()
        if not b_id:
            messagebox.showerror("Error", "Enter Beautician ID to delete.")
            return
        
        confirm = messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete Beautician ID: {b_id}?")
        if not confirm:
            return
            
        try:
            ensure_workbook()
            wb = load_workbook(FILE_NAME)
            ws = wb[SHEET_NAME]
            
            r = find_row_by_army_no(ws, b_id)
            if r is not None:
                ws.delete_rows(r, 1)
                wb.save(FILE_NAME)
                
                # Delete attendance records
                if os.path.exists(ATTENDANCE_FILE):
                    try:
                        awb = load_workbook(ATTENDANCE_FILE)
                        aws = awb.active
                        for ar in range(aws.max_row, 1, -1):
                            if str(aws.cell(row=ar, column=3).value or "").strip() == b_id:
                                aws.delete_rows(ar, 1)
                        # Re-index the S.no column (column 1)
                        for ar in range(2, aws.max_row + 1):
                            aws.cell(row=ar, column=1, value=ar-1)
                        awb.save(ATTENDANCE_FILE)
                        awb.close()
                    except Exception as ae:
                        print("Error deleting attendance records:", ae)

                # Delete service transactions log records
                if os.path.exists(SERVICES_LOG_FILE):
                    try:
                        swb = load_workbook(SERVICES_LOG_FILE)
                        sws = swb.active
                        for sr in range(sws.max_row, 1, -1):
                            if str(sws.cell(row=sr, column=4).value or "").strip() == b_id:
                                sws.delete_rows(sr, 1)
                        # Re-index the S.no column (column 1)
                        for sr in range(2, sws.max_row + 1):
                            sws.cell(row=sr, column=1, value=sr-1)
                        swb.save(SERVICES_LOG_FILE)
                        swb.close()
                    except Exception as se:
                        print("Error deleting services log records:", se)

                for e in [".jpg", ".jpeg", ".png"]:
                    p = os.path.join(KNOWN_FACES_DIR, f"{b_id}{e}")
                    if os.path.exists(p):
                        try:
                            os.remove(p)
                        except:
                            pass
                
                messagebox.showinfo("Success", f"Beautician ID: {b_id} deleted successfully (including all attendance and service records).")
                self.inputs["ID"].delete(0, tk.END)
                self.inputs["Name"].delete(0, tk.END)
                self.load_preview()
                
                threading.Thread(target=load_known_faces, daemon=True).start()
            else:
                messagebox.showerror("Error", f"Beautician ID: {b_id} not found.")
            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Could not delete: {e}")

    def open_attendance_window(self):
        AttendanceWindow(self)

    def open_face_recognition(self):
        FaceRecognitionWindow(self)

    def open_rate_list_editor(self):
        RateListWindow(self)

    def exit_app(self):
        self.destroy()

    def reset_database(self):
        confirm1 = messagebox.askyesno(
            "Confirm Reset", 
            "WARNING: This will permanently delete all registered beauticians, attendance logs, and service records.\n\nAre you sure you want to reset the database?"
        )
        if not confirm1:
            return
            
        confirm2 = messagebox.askyesno(
            "Final Warning",
            "This action CANNOT be undone. Are you absolutely sure you want to proceed and erase everything?"
        )
        if not confirm2:
            return
            
        try:
            # Re-create empty workbooks
            # 1. pers_data.xlsx
            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
            for i, h in enumerate(COLUMNS, start=1):
                ws.cell(row=1, column=i, value=h)
            wb.save(FILE_NAME)
            wb.close()
            
            # 2. attendance.xlsx
            wb_att = Workbook()
            ws_att = wb_att.active
            headers_att = ["S.no", "Date", "Beautician ID", "Name", "Check-in Time", "Check-out Time"]
            for i, h in enumerate(headers_att, start=1):
                ws_att.cell(row=1, column=i, value=h)
            wb_att.save(ATTENDANCE_FILE)
            wb_att.close()
            
            # 3. services_log.xlsx
            wb_serv = Workbook()
            ws_serv = wb_serv.active
            headers_serv = ["S.no", "Date", "Time", "Beautician ID", "Beautician Name", "Category", "Service Name", "Rate"]
            for i, h in enumerate(headers_serv, start=1):
                ws_serv.cell(row=1, column=i, value=h)
            wb_serv.save(SERVICES_LOG_FILE)
            wb_serv.close()
            
            # 4. Delete photos in faces_kaya
            if os.path.exists(KNOWN_FACES_DIR):
                for filename in os.listdir(KNOWN_FACES_DIR):
                    p = os.path.join(KNOWN_FACES_DIR, filename)
                    if os.path.isfile(p):
                        try: os.remove(p)
                        except Exception: pass
                        
            # 5. Delete face encodings cache
            if os.path.exists(ENCODINGS_CACHE_FILE):
                try: os.remove(ENCODINGS_CACHE_FILE)
                except Exception: pass
                
            # 6. Reload and refresh
            self.inputs["ID"].delete(0, tk.END)
            self.inputs["Name"].delete(0, tk.END)
            self.load_preview()
            
            threading.Thread(target=load_known_faces, daemon=True).start()
            messagebox.showinfo("Success", "Database has been reset successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to reset database: {e}")

# ----------- Login -----------
class LoginWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Kaya Kalp Login")
        script_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(script_dir, "favicon.ico")
        if os.path.exists(icon_path):
            self.iconbitmap(icon_path)
        self.geometry("360x220")
        self.resizable(False, False)
        self.configure(bg="#fdf4f5")

        tk.Label(self, text="Username", bg="#fdf4f5", font=("Outfit", 10, "bold")).pack(pady=(20,4))
        self.user_entry = ttk.Entry(self, width=28)
        self.user_entry.pack()
        self.user_entry.insert(0, "admin")

        tk.Label(self, text="Password", bg="#fdf4f5", font=("Outfit", 10, "bold")).pack(pady=(14,4))
        self.pass_entry = ttk.Entry(self, width=28, show="*")
        self.pass_entry.pack()
        self.pass_entry.focus_set()

        btn = tk.Button(self, text="Login", command=self.handle_login, bg="#d48a97", fg="white", font=("Outfit", 10, "bold"))
        btn.pack(pady=16)

        self.bind("<Return>", lambda e: self.handle_login())
        
        ensure_workbook()
        threading.Thread(target=load_known_faces, daemon=True).start()

    def handle_login(self):
        uname = self.user_entry.get().strip()
        pwd = self.pass_entry.get()
        if uname != LOGIN_USERNAME or pwd != LOGIN_PASSWORD:
            messagebox.showerror("Login Failed", "Invalid username or password.")
            return
        self.destroy()
        app = App()
        app.mainloop()

# ----------- Helpers -----------
def get_local_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('10.255.255.255', 1))
        IP = s.getsockname()[0]
    except Exception:
        IP = '127.0.0.1'
    finally:
        s.close()
    return IP

def ensure_ssl_cert(cert_file, key_file):
    """Generate a self-signed certificate if one doesn't already exist."""
    if os.path.exists(cert_file) and os.path.exists(key_file):
        return
    try:
        from cryptography import x509
        from cryptography.x509.oid import NameOID
        from cryptography.hazmat.primitives import hashes, serialization
        from cryptography.hazmat.primitives.asymmetric import rsa
        import ipaddress
        import datetime as dt

        local_ip = get_local_ip()
        key = rsa.generate_private_key(public_exponent=65537, key_size=2048)
        subject = issuer = x509.Name([
            x509.NameAttribute(NameOID.COMMON_NAME, local_ip),
        ])
        san_list = [x509.DNSName("localhost")]
        try:
            san_list.append(x509.IPAddress(ipaddress.ip_address(local_ip)))
        except Exception:
            pass
        cert = (
            x509.CertificateBuilder()
            .subject_name(subject)
            .issuer_name(issuer)
            .public_key(key.public_key())
            .serial_number(x509.random_serial_number())
            .not_valid_before(dt.datetime.utcnow())
            .not_valid_after(dt.datetime.utcnow() + dt.timedelta(days=3650))
            .add_extension(
                x509.SubjectAlternativeName(san_list),
                critical=False
            )
            .sign(key, hashes.SHA256())
        )
        with open(key_file, "wb") as f:
            f.write(key.private_bytes(
                serialization.Encoding.PEM,
                serialization.PrivateFormat.TraditionalOpenSSL,
                serialization.NoEncryption()
            ))
        with open(cert_file, "wb") as f:
            f.write(cert.public_bytes(serialization.Encoding.PEM))
        print("SSL certificate generated successfully.")
    except Exception as e:
        print(f"Failed to generate SSL certificate: {e}")
        raise

# ----------- Entry point -----------
if __name__ == "__main__":
    login = LoginWindow()
    login.mainloop()
